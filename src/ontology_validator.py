# prendi dal rispettivo notebook di marco
import os
import uuid

from llama_index.core import PropertyGraphIndex
from llama_index.core.indices.property_graph import SchemaLLMPathExtractor
from llama_index.embeddings.openai import OpenAIEmbedding

from typing import Literal
from llama_index.core.indices.property_graph import SchemaLLMPathExtractor
from llama_index.core import VectorStoreIndex, SimpleDirectoryReader

from llama_index.llms.openai import OpenAI

# Create a temporary folder to store in-memory files in, which is removed after use.
# Takes a dictionary of the form {filename: bytes} so it also works for non-streamlit applications.
# Note that this removes all file metadata.
class TempDir:
    def __init__(self, files):
        # This assumes Unix-like filesystem.
        # Consider using the tempfile module to make it cross-platform
        self.tmpdir = os.path.join("/tmp/upload/", str(uuid.uuid4()))
        self.files = files

    def __enter__(self):
        os.makedirs(self.tmpdir)
        for filename, file_bytes in self.files.items():
            file_path = os.path.join(self.tmpdir, filename)
            with open(file_path, 'wb') as f:
                f.write(file_bytes)
        return self.tmpdir

    def __exit__(self, exc_type, exc_value, exc_traceback):
        for filename in self.files.keys():
            file_path = os.path.join(self.tmpdir, filename)
            os.remove(file_path)
        os.rmdir(self.tmpdir)
        return False
    
from deeponto.onto import Ontology
from neo4j import GraphDatabase


def compute_label_usage_percentage(uri, user, password, reference_labels):
    """
    Computes the percentage of distinct labels in the database that are part of a reference list.

    Args:
        uri (str): The URI of the Neo4j database (e.g., "bolt://localhost:7687").
        user (str): The username for authentication.
        password (str): The password for authentication.
        reference_labels (list): A list of reference labels to compare against.

    Returns:
        float: The percentage of distinct labels used in the database that match the reference list.
        list: The labels found in both the database and the reference list.
    """
    query = """
    MATCH (n)
    WITH DISTINCT labels(n) AS labels
    UNWIND labels AS label
    RETURN DISTINCT label
    """
    try:
        # Initialize the driver
        driver = GraphDatabase.driver(uri, auth=(user, password))
        
        # Run the query and collect results
        with driver.session() as session:
            result = session.run(query)
            db_labels = {record["label"] for record in result}  # Use a set for fast lookups
        
        # Find the intersection of database labels and reference labels
        reference_set = set(reference_labels)
        common_labels = db_labels.intersection(reference_set)
        
        # Calculate the percentage
        percentage = (len(common_labels) / len(reference_set)) * 100 if reference_set else 0
        
        return percentage, list(common_labels)
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return 0, []
    
    finally:
        # Close the driver
        driver.close()



def compute_relationship_type_usage_percentage(uri, user, password, reference_relationship_types):
    """
    Computes the percentage of distinct relationship types in the database that are part of a reference list.

    Args:
        uri (str): The URI of the Neo4j database (e.g., "bolt://localhost:7687").
        user (str): The username for authentication.
        password (str): The password for authentication.
        reference_relationship_types (list): A list of reference relationship types to compare against.

    Returns:
        float: The percentage of distinct relationship types used in the database that match the reference list.
        list: The relationship types found in both the database and the reference list.
    """
    query = """
    MATCH ()-[r]-()
    RETURN DISTINCT type(r) AS relationship_type
    """
    try:
        # Initialize the driver
        driver = GraphDatabase.driver(uri, auth=(user, password))
        
        # Run the query and collect results
        with driver.session() as session:
            result = session.run(query)
            db_relationship_types = {record["relationship_type"] for record in result}  # Use a set for fast lookups
        
        # Find the intersection of database relationship types and reference list
        reference_set = set(reference_relationship_types)
        common_relationship_types = db_relationship_types.intersection(reference_set)
        
        # Calculate the percentage
        percentage = (len(common_relationship_types) / len(reference_set)) * 100 if reference_set else 0
        
        return percentage, list(common_relationship_types)
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return 0, []
    
    finally:
        # Close the driver
        driver.close()


import openpyxl
from openpyxl import load_workbook


def valuta_txt_ent():
    ent, rel = extract_ent_rel()
    uri = "bolt://localhost:7687"
    user = "neo4j"
    password = "hackaton"
    res = compute_label_usage_percentage(uri, user, password, ent)

    return res

def valuta_txt_rel():
    ent, rel = extract_ent_rel()
    uri = "bolt://localhost:7687"
    user = "neo4j"
    password = "hackaton"
    res = compute_relationship_type_usage_percentage(uri, user, password, rel)

    return res
    
    
def docs2neo4j(documents):
    llm = OpenAI("gpt-4o-mini")

    ent, rel = extract_ent_rel_2()

    # recommended uppercase, underscore separated
    entities = Literal[tuple([re.sub(r'(?<!^)(?=[A-Z])', '_', e).upper() for e in ent])]
    relations = Literal[tuple([re.sub(r'(?<!^)(?=[A-Z])', '_', r).upper() for r in rel])]
    
    
    kg_extractor = SchemaLLMPathExtractor(
        llm=llm,
        possible_entities=entities,
        possible_relations=relations,
        #kg_validation_schema=schema,
        strict=False,  # if false, will allow triplets outside of the schema
        num_workers=4,
        max_triplets_per_chunk=10,
    )

    # Note: used to be `Neo4jPGStore`
    graph_store = Neo4jPropertyGraphStore(
        username="neo4j",
        password="hackaton",
        url="bolt://localhost:7687",
    )

    index = PropertyGraphIndex.from_documents(
        documents,
        kg_extractors=[kg_extractor],
        property_graph_store=graph_store,
        show_progress=True,
    )

    


def pialla_neo4j():
    query = """
        MATCH (n)
        DETACH DELETE (n)
        """
    try:
        # Initialize the driver
        driver = GraphDatabase.driver(uri, auth=(user, password))
        
        # Run the query and collect results
        with driver.session() as session:
            result = session.run(query)
    except Exception as e:
        print(f"An error occurred: {e}")
            
import re

def txt2doc(txt):
    file_dict = {"test": open(txt, "r", errors='ignore').read().encode()}
    with TempDir(file_dict) as tempdir:
        documents = SimpleDirectoryReader(input_dir=tempdir).load_data()
    return documents

def extract_ent_rel():
    onto = Ontology("/Users/marcocalamo/multi_ont.owl", "elk")
    ent = []
    for s,_ in onto.owl_classes.items():
        #print(s.split('#')[1])
        ent.append(s.split('#')[1])
        

    rel = []
    for s,_ in onto.owl_object_properties.items():
    #    #print(s.split('#')[1])
        rel.append(s.split('#')[1])

    entities = [re.sub(r'(?<!^)(?=[A-Z])', '_', e).upper() for e in ent]
    relations = [re.sub(r'(?<!^)(?=[A-Z])', '_', r).upper() for r in rel]

    #print(entities, relations)

    return entities, relations
    

def genera_excel_con_valutazione(input_txt, output_excel):
    """
    Legge un file txt, chiama una funzione di valutazione e salva i risultati in un file Excel.
    
    Args:
        input_txt (str): Percorso del file .txt da valutare.
        output_excel (str): Percorso del file .xlsx generato.
    """

    nome_txt = input_txt.split("/")[-1]  # Estrae il nome del file

    # Crea doc
    document = txt2doc(input_txt)
    
    # Riempi neo4j
    docs2neo4j(document)
    
    # Ottieni la valutazione e le entità trovate
    valutazione, entita_trovate = valuta_txt_ent()
    
    # Tronca la valutazione al secondo decimale
    valutazione_troncata = round(valutazione, 2)
    
    # Prepara i dati da salvare
    entita_str = ", ".join(entita_trovate)  # Formatta le entità come stringa

    # Ottieni la valutazione e le entità trovate
    valutazione_rel, rel_trovate = valuta_txt_rel()
    
    # Tronca la valutazione al secondo decimale
    valutazione_r_troncata = round(valutazione_rel, 2)
    
    # Prepara i dati da salvare
    rel_str = ", ".join(rel_trovate)  # Formatta le entità come stringa
    
   # Verifica se il file Excel esiste
    if os.path.exists(output_excel):
        # Carica il file esistente
        wb = load_workbook(output_excel)
        ws = wb.active
    else:
        # Crea un nuovo file Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valutazioni"
        # Scrivi l'intestazione
        ws.append(["nome_txt", "valutazione_ent", "entita_trovate", "valutazione_rel", "rel_trovate"])
    
    # Scrivi i dati
    ws.append([nome_txt, valutazione_troncata, entita_str, valutazione_r_troncata, rel_str])
    
    # Salva il file Excel
    wb.save(output_excel)
    print(f"File Excel salvato in: {output_excel}")

    pialla_neo4j()
